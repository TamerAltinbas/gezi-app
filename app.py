import os
import io
import csv
import sqlite3
from datetime import datetime, timedelta
from functools import wraps
from openpyxl import load_workbook
from flask import Flask, request, redirect, url_for, render_template_string, session, flash, send_file, g

APP_TITLE = "ÇANAKKALE - İSTANBUL GEZİSİ"
TRIP_DATES = "16-17-18-19 Mayıs 2026"
APPLY_DEADLINE = datetime(2026, 5, 1, 23, 59)
CANCEL_DEADLINE = datetime(2026, 5, 5, 23, 59)
TOTAL_QUOTA = 150
TRIP_FEE_TEXT = "Ücret ve ödeme takvimi okul yönetimi tarafından ayrıca duyurulacaktır."
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "1234")
SECRET_KEY = os.environ.get("SECRET_KEY", "degistir-bu-anahtari")
DB_PATH = os.path.join(os.path.dirname(__file__), "gezi.db")
DEFAULT_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "8. SINIF ÖĞRENCİLERİ.xlsx")

app = Flask(__name__)
app.secret_key = SECRET_KEY


def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(error=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH)
    cur = db.cursor()
    cur.executescript(
        """
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_name TEXT,
            school_no TEXT NOT NULL,
            name TEXT NOT NULL,
            tc TEXT NOT NULL,
            student_group TEXT,
            active INTEGER DEFAULT 1,
            created_at TEXT NOT NULL,
            UNIQUE(tc, school_no)
        );

        CREATE TABLE IF NOT EXISTS applications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            app_no TEXT UNIQUE NOT NULL,
            class_name TEXT,
            school_no TEXT NOT NULL,
            name TEXT NOT NULL,
            tc TEXT NOT NULL,
            student_group TEXT,
            status TEXT NOT NULL,
            payment_status TEXT NOT NULL DEFAULT 'BEKLIYOR',
            payment_amount TEXT DEFAULT '',
            payment_note TEXT DEFAULT '',
            extra_until TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            canceled_at TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT NOT NULL,
            detail TEXT NOT NULL,
            created_at TEXT NOT NULL,
            ip_address TEXT DEFAULT ''
        );
        """
    )
    defaults = {
        'total_quota': str(TOTAL_QUOTA),
        'apply_deadline': APPLY_DEADLINE.isoformat(timespec='minutes'),
        'cancel_deadline': CANCEL_DEADLINE.isoformat(timespec='minutes'),
    }
    for k, v in defaults.items():
        cur.execute("INSERT OR IGNORE INTO settings(key, value) VALUES (?, ?)", (k, v))
    db.commit()
    db.close()


def log_action(action, detail):
    db = get_db()
    db.execute(
        "INSERT INTO logs(action, detail, created_at, ip_address) VALUES (?, ?, ?, ?)",
        (action, detail, datetime.now().isoformat(timespec='seconds'), request.headers.get('X-Forwarded-For', request.remote_addr or '')),
    )
    db.commit()


def get_setting(key, default=None):
    row = get_db().execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return row['value'] if row else default


def set_setting(key, value):
    db = get_db()
    db.execute("INSERT INTO settings(key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (key, value))
    db.commit()


def dt_from_setting(key, default_dt):
    raw = get_setting(key, default_dt.isoformat(timespec='minutes'))
    try:
        return datetime.fromisoformat(raw)
    except Exception:
        return default_dt


def total_quota():
    try:
        return int(get_setting('total_quota', str(TOTAL_QUOTA)))
    except Exception:
        return TOTAL_QUOTA


def active_count():
    row = get_db().execute("SELECT COUNT(*) AS c FROM applications WHERE status='AKTIF'").fetchone()
    return int(row['c'])


def remaining_quota():
    return max(0, total_quota() - active_count())


def app_closed():
    return datetime.now() > dt_from_setting('apply_deadline', APPLY_DEADLINE) or remaining_quota() <= 0


def fmt_dt(value):
    if not value:
        return '-'
    try:
        dt = datetime.fromisoformat(value)
    except Exception:
        return value
    months = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
    return f"{dt.day:02d} {months[dt.month-1]} {dt.year} {dt.hour:02d}:{dt.minute:02d}"


def countdown_text(target_dt):
    diff = target_dt - datetime.now()
    if diff.total_seconds() <= 0:
        return "Süre doldu"
    total_seconds = int(diff.total_seconds())
    days = total_seconds // 86400
    hours = (total_seconds % 86400) // 3600
    minutes = (total_seconds % 3600) // 60
    return f"{days} gün {hours} saat {minutes} dakika"


def next_app_no():
    row = get_db().execute("SELECT COUNT(*) AS c FROM applications").fetchone()
    return f"CZ-{1001 + int(row['c']) - 0}"


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get('admin_ok'):
            return redirect(url_for('admin_login'))
        return fn(*args, **kwargs)
    return wrapper


def import_students_from_excel(path):
    wb = load_workbook(path, data_only=True)
    if 'Sayfa2' not in wb.sheetnames:
        raise ValueError("Excel dosyasında 'Sayfa2' bulunamadı.")
    ws = wb['Sayfa2']
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    header = [str(x).strip() if x is not None else '' for x in rows[0]]
    expected = ['SINIF', 'ÖĞRENCİ NO', 'ÖĞRENCİ ADI SOYADI', 'TC. KİMLİK NO', 'ÖĞRENCİ GRUBU']
    if header[:5] != expected:
        raise ValueError("Excel şablonu beklenen başlıklara uymuyor.")

    db = get_db()
    inserted = 0
    db.execute("DELETE FROM students")
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == '' for v in row[:5]):
            continue
        class_name = str(row[0]).strip() if row[0] is not None else ''
        school_no = str(row[1]).strip() if row[1] is not None else ''
        name = str(row[2]).strip() if row[2] is not None else ''
        tc_raw = str(row[3]).strip() if row[3] is not None else ''
        tc = ''.join(ch for ch in tc_raw if ch.isdigit())
        group = str(row[4]).strip() if row[4] is not None else ''
        if not (school_no and name and tc):
            continue
        db.execute(
            "INSERT OR REPLACE INTO students(class_name, school_no, name, tc, student_group, active, created_at) VALUES (?, ?, ?, ?, ?, 1, ?)",
            (class_name, school_no, name, tc, group, datetime.now().isoformat(timespec='seconds')),
        )
        inserted += 1
    db.commit()
    return inserted


def base_context():
    apply_deadline = dt_from_setting('apply_deadline', APPLY_DEADLINE)
    cancel_deadline = dt_from_setting('cancel_deadline', CANCEL_DEADLINE)
    db = get_db()
    student_count = db.execute("SELECT COUNT(*) AS c FROM students WHERE active=1").fetchone()['c']
    latest_app = db.execute("SELECT app_no, created_at FROM applications ORDER BY id DESC LIMIT 1").fetchone()
    return {
        'app_title': APP_TITLE,
        'trip_dates': TRIP_DATES,
        'apply_deadline': apply_deadline,
        'cancel_deadline': cancel_deadline,
        'apply_deadline_text': fmt_dt(apply_deadline.isoformat(timespec='minutes')),
        'cancel_deadline_text': fmt_dt(cancel_deadline.isoformat(timespec='minutes')),
        'countdown_text': countdown_text(apply_deadline),
        'total_quota': total_quota(),
        'active_count': active_count(),
        'remaining_quota': remaining_quota(),
        'is_closed': app_closed(),
        'student_count': student_count,
        'latest_app': latest_app,
        'trip_fee_text': TRIP_FEE_TEXT,
    }


BASE_HTML = """
<!doctype html>
<html lang="tr">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{{ title or app_title }}</title>
  <style>
    *{box-sizing:border-box} body{margin:0;font-family:Arial,Helvetica,sans-serif;background:#f5f7fb;color:#1f2937}
    a{text-decoration:none;color:inherit}.wrap{max-width:1180px;margin:0 auto;padding:24px}
    .top{display:flex;justify-content:space-between;align-items:center;gap:12px;background:#fff;border:1px solid #e5e7eb;border-radius:24px;padding:24px;box-shadow:0 8px 24px rgba(15,23,42,.05)}
    .badge{display:inline-block;background:#eff6ff;color:#1d4ed8;border:1px solid #bfdbfe;padding:6px 12px;border-radius:999px;font-size:12px;font-weight:700;margin-right:8px}
    .h1{font-size:32px;font-weight:900;margin:6px 0 4px}.muted{color:#6b7280}.grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;margin-top:20px}.card{background:#fff;border:1px solid #e5e7eb;border-radius:24px;box-shadow:0 8px 24px rgba(15,23,42,.04)}
    .stat{padding:20px}.stat .label{font-size:13px;color:#6b7280}.stat .value{margin-top:6px;font-size:28px;font-weight:800}.main{display:grid;grid-template-columns:1.3fr .9fr;gap:20px;margin-top:20px}.body{padding:22px}.btn{display:inline-block;border:none;border-radius:16px;padding:12px 18px;font-weight:700;cursor:pointer;background:#111827;color:#fff}.btn.alt{background:#fff;color:#111827;border:1px solid #d1d5db}.btn.red{background:#b91c1c}.btn.green{background:#166534}.btn.small{padding:9px 14px;border-radius:12px;font-size:14px}
    .row{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}.field{display:flex;flex-direction:column;gap:8px}.field label{font-size:14px;font-weight:700}.field input,.field textarea{width:100%;padding:12px 14px;border:1px solid #d1d5db;border-radius:14px;font-size:14px}.field textarea{min-height:120px}
    .alert{padding:16px;border-radius:18px;border:1px solid #dbeafe;background:#eff6ff;color:#1e3a8a}.warn{border-color:#fde68a;background:#fffbeb;color:#92400e}.ok{border-color:#bbf7d0;background:#f0fdf4;color:#166534}
    .nav{display:flex;flex-wrap:wrap;gap:10px;margin-top:18px}.nav a{background:#fff;border:1px solid #d1d5db;padding:10px 14px;border-radius:14px;font-weight:700}.flash{margin:14px 0;padding:14px 16px;background:#fff7ed;border:1px solid #fed7aa;border-radius:16px;color:#9a3412}
    table{width:100%;border-collapse:collapse} th,td{padding:12px 10px;border-bottom:1px solid #e5e7eb;text-align:left;font-size:14px} th{background:#f8fafc;position:sticky;top:0} .table-wrap{max-height:420px;overflow:auto;border:1px solid #e5e7eb;border-radius:18px}
    .split{display:grid;grid-template-columns:1fr 1fr;gap:20px}.kv{display:grid;grid-template-columns:180px 1fr;gap:8px 12px;font-size:14px;line-height:1.7}.footer-note{font-size:12px;color:#6b7280;margin-top:12px}
    @media (max-width:900px){.grid3,.main,.row,.split{grid-template-columns:1fr}.top{align-items:flex-start;flex-direction:column}.h1{font-size:26px}.kv{grid-template-columns:1fr}}
  </style>
</head>
<body>
<div class="wrap">
  <div class="top">
    <div>
      <div><span class="badge">Güvenli Başvuru Sistemi</span><span class="badge">Excel Şablonu Destekli</span></div>
      <div class="h1">{{ app_title }}</div>
      <div class="muted">{{ trip_dates }}</div>
    </div>
    <div class="nav">
      <a href="{{ url_for('home') }}">Ana Sayfa</a>
      <a href="{{ url_for('apply') }}">Başvuru</a>
      <a href="{{ url_for('status_query') }}">Durum Sorgula</a>
      <a href="{{ url_for('cancel_application') }}">Başvuru İptal</a>
      <a href="{{ url_for('admin_login') }}">Admin</a>
    </div>
  </div>

  {% with messages = get_flashed_messages() %}
    {% if messages %}
      {% for message in messages %}<div class="flash">{{ message }}</div>{% endfor %}
    {% endif %}
  {% endwith %}

  {{ body|safe }}
</div>
</body>
</html>
"""


def render_page(body, **ctx):
    full = base_context()
    full.update(ctx)
    return render_template_string(BASE_HTML, body=body, **full)


@app.route('/')
def home():
    ctx = base_context()
    body = render_template_string(
        """
        <div class="grid3">
          <div class="card stat"><div class="label">Başvurunun bitmesine</div><div class="value">{{ countdown_text }}</div><div class="muted">Son başvuru: {{ apply_deadline_text }}</div></div>
          <div class="card stat"><div class="label">Boş kontenjan</div><div class="value">{{ remaining_quota }} öğrenci</div><div class="muted">Toplam kontenjan: {{ total_quota }}</div></div>
          <div class="card stat"><div class="label">Başvuru durumu</div><div class="value">{{ 'Kapalı' if is_closed else 'Devam ediyor' }}</div><div class="muted">İptal son tarihi: {{ cancel_deadline_text }}</div></div>
        </div>
        <div class="main">
          <div class="card body">
            <h2>Gezi Başvuru Bilgilendirmesi</h2>
            <p>Veliler için resmi başvuru ekranıdır. Sahte veya mükerrer başvuruyu önlemek amacıyla sistem, yönetim panelinden yüklenen öğrenci verileri üzerinden çalışır.</p>
            <div class="alert">Doğrulama yöntemi: <strong>TC Kimlik No + Okul No</strong>. Her öğrenci için yalnızca bir aktif başvuru alınır.</div>
            <div class="split" style="margin-top:16px;">
              <div class="card body"><strong>Başvuru bitiş tarihi</strong><div style="margin-top:8px;font-size:20px;font-weight:800">{{ apply_deadline_text }}</div></div>
              <div class="card body"><strong>İptal son tarihi</strong><div style="margin-top:8px;font-size:20px;font-weight:800">{{ cancel_deadline_text }}</div></div>
            </div>
            <div style="margin-top:16px;display:flex;gap:10px;flex-wrap:wrap;">
              <a class="btn" href="{{ url_for('apply') }}">Başvuruya Başla</a>
              <a class="btn alt" href="{{ url_for('status_query') }}">Durum Sorgula</a>
              <a class="btn alt" href="{{ url_for('cancel_application') }}">Başvuru İptal Et</a>
            </div>
          </div>
          <div class="card body">
            <h2>Canlı Özet</h2>
            <div class="kv">
              <div>Toplam kontenjan</div><div><strong>{{ total_quota }}</strong></div>
              <div>Aktif başvuru</div><div><strong>{{ active_count }}</strong></div>
              <div>Boş kontenjan</div><div><strong>{{ remaining_quota }}</strong></div>
              <div>Yüklü öğrenci</div><div><strong>{{ student_count }}</strong></div>
              <div>Son başvuru</div><div><strong>{{ latest_app['app_no'] if latest_app else 'Henüz yok' }}</strong>{% if latest_app %} – {{ latest_app['created_at'][:16].replace('T',' ') }}{% endif %}</div>
            </div>
            <div class="footer-note">Not: İptal edilen başvurular kontenjana tekrar eklenir.</div>
          </div>
        </div>
        """,
        **ctx,
    )
    return render_page(body)


@app.route('/apply', methods=['GET', 'POST'])
def apply():
    db = get_db()
    student = None
    if request.method == 'POST':
        tc = ''.join(ch for ch in request.form.get('tc', '').strip() if ch.isdigit())
        school_no = request.form.get('school_no', '').strip()
        if app_closed():
            flash('Başvuru kapalıdır. Süre dolmuş olabilir veya kontenjan bitmiş olabilir.')
            return redirect(url_for('apply'))
        student = db.execute("SELECT * FROM students WHERE tc=? AND school_no=? AND active=1", (tc, school_no)).fetchone()
        if not student:
            flash('Öğrenci bilgileri doğrulanamadı.')
            return redirect(url_for('apply'))
        existing = db.execute("SELECT * FROM applications WHERE tc=? AND school_no=? AND status='AKTIF'", (tc, school_no)).fetchone()
        if existing:
            flash(f"Bu öğrenci için zaten aktif başvuru var. Başvuru No: {existing['app_no']}")
            return redirect(url_for('status_query', tc=tc, school_no=school_no, app_no=existing['app_no']))

        checks = [request.form.get('check1'), request.form.get('check2'), request.form.get('check3'), request.form.get('check4')]
        if request.form.get('confirm') == '1':
            if not all(checks):
                flash('Lütfen tüm onay kutularını işaretleyin.')
                return redirect(url_for('apply'))
            app_no = next_app_no()
            db.execute(
                "INSERT INTO applications(app_no, class_name, school_no, name, tc, student_group, status, created_at) VALUES (?, ?, ?, ?, ?, ?, 'AKTIF', ?)",
                (app_no, student['class_name'], student['school_no'], student['name'], student['tc'], student['student_group'], datetime.now().isoformat(timespec='seconds')),
            )
            db.commit()
            log_action('BASVURU_OLUSTURULDU', f"{app_no} - {student['name']}")
            flash(f"Başvurunuz kaydedildi. Başvuru Numaranız: {app_no}")
            return redirect(url_for('status_query', tc=student['tc'], school_no=student['school_no'], app_no=app_no))

    body = render_template_string(
        """
        <div class="card body">
          <h2>Veli Başvuru Ekranı</h2>
          {% if is_closed %}<div class="alert warn">Başvuru şu anda kapalı. Yeni başvuru alınmıyor.</div>{% endif %}
          <form method="post">
            <div class="row">
              <div class="field"><label>Öğrenci TC Kimlik No</label><input name="tc" maxlength="11" required></div>
              <div class="field"><label>Okul Numarası</label><input name="school_no" required></div>
              <div class="field" style="align-self:end"><button class="btn" type="submit" {% if is_closed %}disabled{% endif %}>Öğrenciyi Doğrula</button></div>
            </div>
            <div class="footer-note">Sistem doğrulaması, admin panelinden yüklenen Excel öğrenci listesi üzerinden yapılır.</div>

            {% if student %}
              <div class="split" style="margin-top:18px;">
                <div class="card body">
                  <h3>Öğrenci Bilgisi</h3>
                  <div class="kv">
                    <div>Öğrenci</div><div><strong>{{ student['name'] }}</strong></div>
                    <div>Sınıf</div><div><strong>{{ student['class_name'] }}</strong></div>
                    <div>Okul No</div><div><strong>{{ student['school_no'] }}</strong></div>
                    <div>Grup</div><div><strong>{{ student['student_group'] or '-' }}</strong></div>
                  </div>
                </div>
                <div class="card body">
                  <h3>Gezi Bilgisi</h3>
                  <div class="kv">
                    <div>Gezi</div><div><strong>{{ app_title }}</strong></div>
                    <div>Tarih</div><div><strong>{{ trip_dates }}</strong></div>
                    <div>Ödeme</div><div>{{ trip_fee_text }}</div>
                  </div>
                </div>
              </div>
              <div class="card body" style="margin-top:18px;">
                <h3>Onaylar</h3>
                <div style="display:grid;gap:10px;">
                  <label><input type="checkbox" name="check1" value="1"> Gezi kurallarını okudum ve kabul ediyorum.</label>
                  <label><input type="checkbox" name="check2" value="1"> Ödeme şartlarını okudum ve kabul ediyorum.</label>
                  <label><input type="checkbox" name="check3" value="1"> İptal şartlarını okudum ve kabul ediyorum.</label>
                  <label><input type="checkbox" name="check4" value="1"> Okul tarafından yapılacak resmi duyuruları takip edeceğimi kabul ediyorum.</label>
                </div>
                <input type="hidden" name="tc" value="{{ student['tc'] }}">
                <input type="hidden" name="school_no" value="{{ student['school_no'] }}">
                <input type="hidden" name="confirm" value="1">
                <div style="margin-top:16px;"><button class="btn green" type="submit">Başvuruyu Kaydet</button></div>
              </div>
            {% endif %}
          </form>
        </div>
        """,
        student=student,
        **base_context(),
    )
    return render_page(body)


@app.route('/status', methods=['GET', 'POST'])
def status_query():
    db = get_db()
    query = {
        'tc': request.values.get('tc', '').strip(),
        'school_no': request.values.get('school_no', '').strip(),
        'app_no': request.values.get('app_no', '').strip(),
    }
    found = None
    if all(query.values()):
        tc = ''.join(ch for ch in query['tc'] if ch.isdigit())
        found = db.execute("SELECT * FROM applications WHERE tc=? AND school_no=? AND app_no=?", (tc, query['school_no'], query['app_no'])).fetchone()
    body = render_template_string(
        """
        <div class="card body">
          <h2>Başvuru Durumu Sorgulama</h2>
          <form method="post">
            <div class="row">
              <div class="field"><label>TC Kimlik No</label><input name="tc" value="{{ query.tc }}" required></div>
              <div class="field"><label>Okul No</label><input name="school_no" value="{{ query.school_no }}" required></div>
              <div class="field"><label>Başvuru No</label><input name="app_no" value="{{ query.app_no }}" required></div>
            </div>
            <div style="margin-top:16px;"><button class="btn" type="submit">Sorgula</button></div>
          </form>
          {% if found %}
            <div class="card body" style="margin-top:18px;">
              <div class="kv">
                <div>Başvuru No</div><div><strong>{{ found['app_no'] }}</strong></div>
                <div>Öğrenci</div><div><strong>{{ found['name'] }}</strong></div>
                <div>Sınıf</div><div><strong>{{ found['class_name'] }}</strong></div>
                <div>Durum</div><div><strong>{{ found['status'] }}</strong></div>
                <div>Ödeme</div><div><strong>{{ found['payment_status'] }}</strong>{% if found['payment_amount'] %} – {{ found['payment_amount'] }}{% endif %}</div>
                <div>Ek Süre</div><div>{{ fmt_dt(found['extra_until']) }}</div>
                <div>Not</div><div>{{ found['payment_note'] or '-' }}</div>
                <div>Başvuru Tarihi</div><div>{{ fmt_dt(found['created_at']) }}</div>
              </div>
            </div>
          {% elif query.tc or query.school_no or query.app_no %}
            <div class="alert warn" style="margin-top:18px;">Girilen bilgilere ait başvuru bulunamadı.</div>
          {% endif %}
        </div>
        """,
        query=query,
        found=found,
        fmt_dt=fmt_dt,
    )
    return render_page(body)


@app.route('/cancel', methods=['GET', 'POST'])
def cancel_application():
    db = get_db()
    found = None
    query = {'tc': '', 'school_no': '', 'app_no': ''}
    if request.method == 'POST':
        query = {
            'tc': request.form.get('tc', '').strip(),
            'school_no': request.form.get('school_no', '').strip(),
            'app_no': request.form.get('app_no', '').strip(),
        }
        tc = ''.join(ch for ch in query['tc'] if ch.isdigit())
        found = db.execute("SELECT * FROM applications WHERE tc=? AND school_no=? AND app_no=?", (tc, query['school_no'], query['app_no'])).fetchone()
        if request.form.get('do_cancel') == '1':
            if not found:
                flash('Başvuru bulunamadı.')
                return redirect(url_for('cancel_application'))
            if datetime.now() > dt_from_setting('cancel_deadline', CANCEL_DEADLINE):
                flash('İptal süresi dolmuştur.')
                return redirect(url_for('cancel_application'))
            if found['status'] == 'IPTAL':
                flash('Bu başvuru zaten iptal edilmiş.')
                return redirect(url_for('cancel_application'))
            db.execute("UPDATE applications SET status='IPTAL', canceled_at=? WHERE id=?", (datetime.now().isoformat(timespec='seconds'), found['id']))
            db.commit()
            log_action('BASVURU_IPTAL', f"{found['app_no']} - {found['name']}")
            flash('Başvuru iptal edildi.')
            return redirect(url_for('status_query', tc=found['tc'], school_no=found['school_no'], app_no=found['app_no']))

    body = render_template_string(
        """
        <div class="card body">
          <h2>Başvuru İptal Ekranı</h2>
          <div class="alert warn">İptal son tarihi: <strong>{{ cancel_deadline_text }}</strong></div>
          <form method="post" style="margin-top:16px;">
            <div class="row">
              <div class="field"><label>TC Kimlik No</label><input name="tc" value="{{ query.tc }}" required></div>
              <div class="field"><label>Okul No</label><input name="school_no" value="{{ query.school_no }}" required></div>
              <div class="field"><label>Başvuru No</label><input name="app_no" value="{{ query.app_no }}" required></div>
            </div>
            <div style="margin-top:16px;"><button class="btn" type="submit">Başvuruyu Bul</button></div>
            {% if found %}
              <div class="card body" style="margin-top:18px;">
                <div class="kv">
                  <div>Başvuru No</div><div><strong>{{ found['app_no'] }}</strong></div>
                  <div>Öğrenci</div><div><strong>{{ found['name'] }}</strong></div>
                  <div>Durum</div><div><strong>{{ found['status'] }}</strong></div>
                </div>
                <input type="hidden" name="tc" value="{{ found['tc'] }}">
                <input type="hidden" name="school_no" value="{{ found['school_no'] }}">
                <input type="hidden" name="app_no" value="{{ found['app_no'] }}">
                <input type="hidden" name="do_cancel" value="1">
                <div style="margin-top:16px;"><button class="btn red" type="submit" {% if found['status']=='IPTAL' %}disabled{% endif %}>Başvuruyu İptal Et</button></div>
              </div>
            {% endif %}
          </form>
        </div>
        """,
        query=query,
        found=found,
        cancel_deadline_text=fmt_dt(dt_from_setting('cancel_deadline', CANCEL_DEADLINE).isoformat(timespec='minutes')),
    )
    return render_page(body)


@app.route('/admin', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        if request.form.get('password') == ADMIN_PASSWORD:
            session['admin_ok'] = True
            return redirect(url_for('admin_dashboard'))
        flash('Admin şifresi hatalı.')
    body = """
    <div class='card body' style='max-width:520px;margin:20px auto;'>
      <h2>Admin Girişi</h2>
      <form method='post'>
        <div class='field'><label>Şifre</label><input type='password' name='password' required></div>
        <div style='margin-top:16px;'><button class='btn' type='submit'>Giriş Yap</button></div>
      </form>
    </div>
    """
    return render_page(body, title='Admin Girişi')


@app.route('/admin/logout')
@admin_required
def admin_logout():
    session.clear()
    flash('Admin oturumu kapatıldı.')
    return redirect(url_for('home'))


@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    db = get_db()
    students = db.execute("SELECT COUNT(*) AS c FROM students WHERE active=1").fetchone()['c']
    apps = db.execute("SELECT * FROM applications ORDER BY id DESC").fetchall()
    logs = db.execute("SELECT * FROM logs ORDER BY id DESC LIMIT 100").fetchall()
    body = render_template_string(
        """
        <div class="grid3">
          <div class="card stat"><div class="label">Yüklü öğrenci</div><div class="value">{{ students }}</div></div>
          <div class="card stat"><div class="label">Toplam başvuru</div><div class="value">{{ apps|length }}</div></div>
          <div class="card stat"><div class="label">Ödeme bekleyen</div><div class="value">{{ apps|selectattr('payment_status','equalto','BEKLIYOR')|list|length }}</div></div>
        </div>
        <div class="nav" style="margin-top:18px;">
          <a href="{{ url_for('admin_import_students') }}">Excel ile Öğrenci Yükle</a>
          <a href="{{ url_for('admin_settings') }}">Ayarlar</a>
          <a href="{{ url_for('admin_payments') }}">Ödeme / Ek Süre</a>
          <a href="{{ url_for('admin_export_applications') }}">Başvuruları CSV İndir</a>
          <a href="{{ url_for('admin_logout') }}">Çıkış</a>
        </div>
        <div class="main">
          <div class="card body">
            <h2>Başvurular</h2>
            <div class="table-wrap">
              <table>
                <thead><tr><th>Başvuru No</th><th>Öğrenci</th><th>Sınıf</th><th>Durum</th><th>Ödeme</th><th>Tarih</th></tr></thead>
                <tbody>
                {% for a in apps %}
                  <tr>
                    <td>{{ a['app_no'] }}</td>
                    <td>{{ a['name'] }}</td>
                    <td>{{ a['class_name'] }}</td>
                    <td>{{ a['status'] }}</td>
                    <td>{{ a['payment_status'] }}{% if a['payment_amount'] %} - {{ a['payment_amount'] }}{% endif %}</td>
                    <td>{{ a['created_at'][:16].replace('T',' ') }}</td>
                  </tr>
                {% endfor %}
                </tbody>
              </table>
            </div>
          </div>
          <div class="card body">
            <h2>İşlem Kayıtları</h2>
            <div class="table-wrap">
              <table>
                <thead><tr><th>İşlem</th><th>Detay</th><th>Zaman</th><th>IP</th></tr></thead>
                <tbody>
                {% for l in logs %}
                  <tr><td>{{ l['action'] }}</td><td>{{ l['detail'] }}</td><td>{{ l['created_at'][:16].replace('T',' ') }}</td><td>{{ l['ip_address'] }}</td></tr>
                {% endfor %}
                </tbody>
              </table>
            </div>
          </div>
        </div>
        """,
        students=students,
        apps=apps,
        logs=logs,
    )
    return render_page(body, title='Admin Paneli')


@app.route('/admin/import', methods=['GET', 'POST'])
@admin_required
def admin_import_students():
    if request.method == 'POST':
        file = request.files.get('excel_file')
        if file and file.filename:
            temp_path = os.path.join(os.path.dirname(__file__), '_temp_import.xlsx')
            file.save(temp_path)
            try:
                count = import_students_from_excel(temp_path)
                log_action('OGRENCI_LISTESI_YUKLENDI', f"{count} öğrenci Excel ile aktarıldı")
                flash(f'{count} öğrenci başarıyla yüklendi.')
            except Exception as e:
                flash(f'Yükleme hatası: {e}')
            finally:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            return redirect(url_for('admin_import_students'))
        elif request.form.get('use_default') == '1':
            if not os.path.exists(DEFAULT_TEMPLATE_PATH):
                flash('Klasörde varsayılan Excel dosyası bulunamadı.')
            else:
                try:
                    count = import_students_from_excel(DEFAULT_TEMPLATE_PATH)
                    log_action('OGRENCI_LISTESI_YUKLENDI', f"{count} öğrenci varsayılan Excel ile aktarıldı")
                    flash(f'{count} öğrenci varsayılan Excel dosyasından yüklendi.')
                except Exception as e:
                    flash(f'Yükleme hatası: {e}')
            return redirect(url_for('admin_import_students'))

    body = render_template_string(
        """
        <div class="card body">
          <h2>Excel ile Öğrenci Yükleme</h2>
          <div class="alert">Beklenen sayfa: <strong>Sayfa2</strong><br>Beklenen sütunlar: <strong>SINIF, ÖĞRENCİ NO, ÖĞRENCİ ADI SOYADI, TC. KİMLİK NO, ÖĞRENCİ GRUBU</strong></div>
          <form method="post" enctype="multipart/form-data" style="margin-top:16px;">
            <div class="field"><label>Excel Dosyası Seç (.xlsx)</label><input type="file" name="excel_file" accept=".xlsx"></div>
            <div style="margin-top:16px;display:flex;gap:10px;flex-wrap:wrap;">
              <button class="btn" type="submit">Excel'i Yükle</button>
            </div>
          </form>
          <form method="post" style="margin-top:12px;">
            <input type="hidden" name="use_default" value="1">
            <button class="btn alt" type="submit">Klasördeki Varsayılan Excel'i Yükle</button>
          </form>
          <div class="footer-note">Yeni öğrenci listesi yüklendiğinde mevcut öğrenci listesi tamamen yenilenir.</div>
        </div>
        """
    )
    return render_page(body, title='Excel ile Öğrenci Yükle')


@app.route('/admin/settings', methods=['GET', 'POST'])
@admin_required
def admin_settings():
    if request.method == 'POST':
        set_setting('total_quota', request.form.get('total_quota', str(TOTAL_QUOTA)).strip())
        set_setting('apply_deadline', request.form.get('apply_deadline').strip())
        set_setting('cancel_deadline', request.form.get('cancel_deadline').strip())
        log_action('AYAR_GUNCELLE', 'Kontenjan ve tarihler güncellendi')
        flash('Ayarlar kaydedildi.')
        return redirect(url_for('admin_settings'))
    body = render_template_string(
        """
        <div class="card body" style="max-width:720px;">
          <h2>Sistem Ayarları</h2>
          <form method="post">
            <div class="row">
              <div class="field"><label>Toplam Kontenjan</label><input name="total_quota" value="{{ total_quota }}"></div>
              <div class="field"><label>Başvuru Son Tarihi</label><input type="datetime-local" name="apply_deadline" value="{{ apply_raw }}"></div>
              <div class="field"><label>İptal Son Tarihi</label><input type="datetime-local" name="cancel_deadline" value="{{ cancel_raw }}"></div>
            </div>
            <div style="margin-top:16px;"><button class="btn" type="submit">Kaydet</button></div>
          </form>
        </div>
        """,
        total_quota=total_quota(),
        apply_raw=dt_from_setting('apply_deadline', APPLY_DEADLINE).strftime('%Y-%m-%dT%H:%M'),
        cancel_raw=dt_from_setting('cancel_deadline', CANCEL_DEADLINE).strftime('%Y-%m-%dT%H:%M'),
    )
    return render_page(body, title='Ayarlar')


@app.route('/admin/payments', methods=['GET', 'POST'])
@admin_required
def admin_payments():
    db = get_db()
    if request.method == 'POST':
        app_no = request.form.get('app_no', '').strip()
        payment_status = request.form.get('payment_status', 'BEKLIYOR').strip()
        payment_amount = request.form.get('payment_amount', '').strip()
        payment_note = request.form.get('payment_note', '').strip()
        extra_days_raw = request.form.get('extra_days', '0').strip() or '0'
        try:
            extra_days = int(extra_days_raw)
        except ValueError:
            extra_days = 0
        extra_until = ''
        if extra_days > 0:
            extra_until = (datetime.now() + timedelta(days=extra_days)).isoformat(timespec='minutes')
        updated = db.execute(
            "UPDATE applications SET payment_status=?, payment_amount=?, payment_note=?, extra_until=? WHERE app_no=?",
            (payment_status, payment_amount, payment_note, extra_until, app_no),
        ).rowcount
        db.commit()
        if updated:
            log_action('ODEME_GUNCELLE', f"{app_no} ödeme:{payment_status} ek_sure:{extra_days} gün")
            flash('Ödeme / ek süre kaydı güncellendi.')
        else:
            flash('Başvuru numarası bulunamadı.')
        return redirect(url_for('admin_payments'))

    apps = db.execute("SELECT * FROM applications ORDER BY id DESC LIMIT 200").fetchall()
    body = render_template_string(
        """
        <div class="split">
          <div class="card body">
            <h2>Ödeme / Ek Süre İşlemi</h2>
            <form method="post">
              <div class="field"><label>Başvuru No</label><input name="app_no" placeholder="CZ-1001" required></div>
              <div class="field"><label>Ödeme Durumu</label>
                <select name="payment_status" style="padding:12px 14px;border:1px solid #d1d5db;border-radius:14px;font-size:14px;">
                  <option value="BEKLIYOR">BEKLIYOR</option>
                  <option value="ALINDI">ALINDI</option>
                  <option value="KISMEN_ALINDI">KISMEN_ALINDI</option>
                </select>
              </div>
              <div class="field"><label>Ödenen Tutar</label><input name="payment_amount" placeholder="8500 TL"></div>
              <div class="field"><label>Ek Süre (gün)</label><input name="extra_days" type="number" min="0" value="0"></div>
              <div class="field"><label>Not</label><textarea name="payment_note"></textarea></div>
              <div style="margin-top:16px;"><button class="btn" type="submit">Kaydet</button></div>
            </form>
          </div>
          <div class="card body">
            <h2>Son Başvurular</h2>
            <div class="table-wrap">
              <table>
                <thead><tr><th>Başvuru</th><th>Öğrenci</th><th>Durum</th><th>Ödeme</th></tr></thead>
                <tbody>
                  {% for a in apps %}
                  <tr><td>{{ a['app_no'] }}</td><td>{{ a['name'] }}</td><td>{{ a['status'] }}</td><td>{{ a['payment_status'] }}</td></tr>
                  {% endfor %}
                </tbody>
              </table>
            </div>
          </div>
        </div>
        """,
        apps=apps,
    )
    return render_page(body, title='Ödeme ve Ek Süre')


@app.route('/admin/export')
@admin_required
def admin_export_applications():
    rows = get_db().execute("SELECT * FROM applications ORDER BY id DESC").fetchall()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Başvuru No', 'Sınıf', 'Okul No', 'Öğrenci', 'TC', 'Grup', 'Durum', 'Ödeme', 'Tutar', 'Not', 'Ek Süre', 'Başvuru Tarihi', 'İptal Tarihi'])
    for r in rows:
        writer.writerow([r['app_no'], r['class_name'], r['school_no'], r['name'], r['tc'], r['student_group'], r['status'], r['payment_status'], r['payment_amount'], r['payment_note'], r['extra_until'], r['created_at'], r['canceled_at']])
    mem = io.BytesIO(output.getvalue().encode('utf-8-sig'))
    mem.seek(0)
    return send_file(mem, as_attachment=True, download_name='gezi_basvurulari.csv', mimetype='text/csv')


if __name__ == '__main__':
    init_db()
    if os.path.exists(DEFAULT_TEMPLATE_PATH):
        with app.app_context():
            db = get_db()
            count = db.execute("SELECT COUNT(*) AS c FROM students").fetchone()['c']
            if count == 0:
                try:
                    imported = import_students_from_excel(DEFAULT_TEMPLATE_PATH)
                    print(f"Varsayılan Excel'den {imported} öğrenci yüklendi.")
                except Exception as e:
                    print(f"Varsayılan Excel yüklenemedi: {e}")
   import os

port = int(os.environ.get("PORT", 5000))
app.run(host="0.0.0.0", port=port)
